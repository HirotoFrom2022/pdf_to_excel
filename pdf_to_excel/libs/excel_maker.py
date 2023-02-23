import openpyxl
import os
import datetime
import pprint

class EXCEL_MAKER:

    def __init__(self, proc):
        self.proc = proc
        self.full_path = ''
        self.loaded_book = ''

    def create_book(self):
        self.proc.log(0, 'Creating excel file.')

        book = openpyxl.Workbook()
        excel_strage_path = self.proc.config.get('STORAGE_PATH', 'excel_folder')
        excel_file_name = datetime.datetime.now().strftime('%Y_%m_%d_%H_%M_%S') + '.xlsx'
        self.full_path = os.path.join(excel_strage_path, excel_file_name)

        book.save(self.full_path)

    def load_book(self):
        self.proc.log(0, 'Loading book.')

        self.loaded_book = openpyxl.load_workbook(self.full_path)

    def write_content_with_p(self, index, request_pages, content):
        self.loaded_book.create_sheet(str(request_pages[index] + 1))
        sheet = self.loaded_book[str(request_pages[index] + 1)]
        for i in range(len(content)):
            # pprint.pprint('---------------------------------DEBUG!!!!-------------------------------------------')
            # pprint.pprint(len(content))
            sheet['A' + str(i+1)] = content[i]

    def write_content_without_p(self, index, content):
        self.loaded_book.create_sheet(str(index + 1))
        sheet = self.loaded_book[str(index + 1)]
        for i in range(len(content)):
            # pprint.pprint('---------------------------------DEBUG!!!!-------------------------------------------')
            # pprint.pprint(len(content))
            sheet['A' + str(i+1)] = content[i]

    def save_book(self):
        self.proc.log(0, 'Saving book.')
        self.loaded_book.remove(self.loaded_book['Sheet'])
        self.loaded_book.save(self.full_path)
